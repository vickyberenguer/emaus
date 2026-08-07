from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy.dialects.mysql import insert as mysql_insert
from pydantic import BaseModel
from datetime import datetime, date
import io
import os
import re
import threading
import unicodedata

from openpyxl import load_workbook

from app.database import get_db
from app.models.usuario import Usuario, RolEnum
from app.models.emaus import Diocesis, Emaus, ResponsableEmaus
from app.models.catalogo import Catalogo
from app.models.establecimiento import EstablecimientoEstado
from app.models.padron_importacion import PadronImportacion
from app.routers.auth import get_current_user, require_rol, hash_password

router = APIRouter(prefix="/admin", tags=["admin"], dependencies=[Depends(require_rol("admin"))])


# ============================================================
# Usuarios
# ============================================================

class UsuarioCreate(BaseModel):
    nombre: str
    apellido: str
    email: str
    password: str
    rol: str  # atl | responsable | admin
    emaus_id: int | None = None


class UsuarioUpdate(BaseModel):
    nombre: str | None = None
    apellido: str | None = None
    email: str | None = None
    password: str | None = None
    rol: str | None = None
    emaus_id: int | None = None
    activo: bool | None = None


class UsuarioResponse(BaseModel):
    id: int
    nombre: str
    apellido: str
    email: str
    rol: str
    activo: bool
    emaus_id: int | None
    creado_en: datetime | None
    ultimo_ingreso: datetime | None

    class Config:
        from_attributes = True


class EmausResponse(BaseModel):
    id: int
    nombre: str
    diocesis_id: int
    diocesis_nombre: str | None = None
    activo: bool

    class Config:
        from_attributes = True


@router.get("/emaus", response_model=list[EmausResponse])
def listar_emaus(db: Session = Depends(get_db)):
    rows = db.query(Emaus, Diocesis.nombre).join(Diocesis, Emaus.diocesis_id == Diocesis.id).order_by(Emaus.nombre).all()
    return [
        EmausResponse(id=e.id, nombre=e.nombre, diocesis_id=e.diocesis_id, diocesis_nombre=d_nombre, activo=e.activo)
        for e, d_nombre in rows
    ]


@router.get("/usuarios", response_model=list[UsuarioResponse])
def listar_usuarios(rol: str | None = None, db: Session = Depends(get_db)):
    q = db.query(Usuario)
    if rol:
        q = q.filter(Usuario.rol == rol)
    return q.order_by(Usuario.apellido, Usuario.nombre).all()


@router.post("/usuarios", response_model=UsuarioResponse, status_code=status.HTTP_201_CREATED)
def crear_usuario(body: UsuarioCreate, db: Session = Depends(get_db)):
    if db.query(Usuario).filter(Usuario.email == body.email.lower().strip()).first():
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Ya existe un usuario con ese email")

    usuario = Usuario(
        nombre=body.nombre,
        apellido=body.apellido,
        email=body.email.lower().strip(),
        password_hash=hash_password(body.password),
        rol=RolEnum(body.rol),
        emaus_id=body.emaus_id,
        activo=True,
    )
    db.add(usuario)
    db.commit()
    db.refresh(usuario)
    return usuario


@router.put("/usuarios/{usuario_id}", response_model=UsuarioResponse)
def actualizar_usuario(usuario_id: int, body: UsuarioUpdate, db: Session = Depends(get_db)):
    usuario = db.query(Usuario).filter(Usuario.id == usuario_id).first()
    if not usuario:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Usuario no encontrado")

    data = body.model_dump(exclude_unset=True)
    if "password" in data:
        usuario.password_hash = hash_password(data.pop("password"))
    if "rol" in data:
        usuario.rol = RolEnum(data.pop("rol"))
    if "email" in data and data["email"]:
        data["email"] = data["email"].lower().strip()

    for field, value in data.items():
        setattr(usuario, field, value)

    db.commit()
    db.refresh(usuario)
    return usuario


# ============================================================
# Asignación Responsable → Emaús
# ============================================================

class ResponsableEmausResponse(BaseModel):
    responsable_id: int
    nombre: str
    apellido: str
    email: str
    emaus_ids: list[int]


class ResponsableEmausUpdate(BaseModel):
    emaus_ids: list[int]


@router.get("/responsable-emaus", response_model=list[ResponsableEmausResponse])
def listar_asignaciones(db: Session = Depends(get_db)):
    responsables = db.query(Usuario).filter(Usuario.rol == RolEnum.responsable).all()
    resultado = []
    for r in responsables:
        emaus_ids = [
            row[0] for row in db.query(ResponsableEmaus.emaus_id)
            .filter(ResponsableEmaus.responsable_id == r.id).all()
        ]
        resultado.append(ResponsableEmausResponse(
            responsable_id=r.id, nombre=r.nombre, apellido=r.apellido, email=r.email, emaus_ids=emaus_ids,
        ))
    return resultado


@router.put("/responsable-emaus/{responsable_id}", response_model=ResponsableEmausResponse)
def actualizar_asignaciones(responsable_id: int, body: ResponsableEmausUpdate, db: Session = Depends(get_db)):
    responsable = db.query(Usuario).filter(
        Usuario.id == responsable_id, Usuario.rol == RolEnum.responsable
    ).first()
    if not responsable:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Responsable no encontrado")

    if body.emaus_ids:
        encontrados = db.query(Emaus.id).filter(Emaus.id.in_(body.emaus_ids)).count()
        if encontrados != len(set(body.emaus_ids)):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Alguno de los Emaús no existe")

    db.query(ResponsableEmaus).filter(ResponsableEmaus.responsable_id == responsable_id).delete()
    for emaus_id in set(body.emaus_ids):
        db.add(ResponsableEmaus(responsable_id=responsable_id, emaus_id=emaus_id))
    db.commit()

    return ResponsableEmausResponse(
        responsable_id=responsable.id, nombre=responsable.nombre, apellido=responsable.apellido,
        email=responsable.email, emaus_ids=list(set(body.emaus_ids)),
    )


# ============================================================
# Catálogos
# ============================================================

class CatalogoCreate(BaseModel):
    categoria: str
    valor: str
    orden: int = 0


class CatalogoUpdate(BaseModel):
    valor: str | None = None
    activo: bool | None = None
    orden: int | None = None


class CatalogoResponse(BaseModel):
    id: int
    categoria: str
    valor: str
    activo: bool
    orden: int

    class Config:
        from_attributes = True


@router.get("/catalogos/{categoria}", response_model=list[CatalogoResponse])
def listar_catalogo(categoria: str, db: Session = Depends(get_db)):
    return db.query(Catalogo).filter(Catalogo.categoria == categoria).order_by(Catalogo.orden).all()


@router.post("/catalogos", response_model=CatalogoResponse, status_code=status.HTTP_201_CREATED)
def crear_item_catalogo(body: CatalogoCreate, db: Session = Depends(get_db)):
    item = Catalogo(categoria=body.categoria, valor=body.valor, orden=body.orden, activo=True)
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.put("/catalogos/{item_id}", response_model=CatalogoResponse)
def actualizar_item_catalogo(item_id: int, body: CatalogoUpdate, db: Session = Depends(get_db)):
    item = db.query(Catalogo).filter(Catalogo.id == item_id).first()
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Ítem de catálogo no encontrado")
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(item, field, value)
    db.commit()
    db.refresh(item)
    return item


# ============================================================
# Padrón de establecimientos educativos
# ============================================================

class PadronEstadoResponse(BaseModel):
    total_registros: int
    ultima_importacion: datetime | None
    ultimo_usuario_id: int | None
    ultimo_total_procesados: int | None
    ultimo_insertados: int | None
    ultimo_actualizados: int | None


# Encabezados esperados en la fila 13 del Excel del Ministerio → columna del modelo.
# Se normalizan (minúsculas, sin acentos) antes de comparar.
ENCABEZADOS_PADRON = {
    "cueanexo": "cueanexo",
    "jurisdiccion": "jurisdiccion",
    "sector": "sector",
    "ambito": "ambito",
    "departamento": "departamento",
    "cod_depto": "cod_departamento",
    "codigo departamento": "cod_departamento",
    "codigo de departamento": "cod_departamento",
    "localidad": "localidad",
    "cod_localidad": "cod_localidad",
    "codigo localidad": "cod_localidad",
    "codigo de localidad": "cod_localidad",
    "nombre": "nombre",
    "domicilio": "domicilio",
    "cp": "codigo_postal",
    "codigo postal": "codigo_postal",
    "c. p.": "codigo_postal",
    "telefono": "telefono",
    "mail": "mail",
    "email": "mail",
    "inicial - jardin maternal": "nivel_inicial_maternal",
    "nivel inicial - jardin maternal": "nivel_inicial_maternal",
    "inicial - jardin de infantes": "nivel_inicial_infantes",
    "nivel inicial - jardin de infantes": "nivel_inicial_infantes",
    "primario": "primario",
    "secundario": "secundario",
    "adultos": "adultos",
    "formacion profesional": "formacion_profesional",
    "alfabetizacion": "alfabetizacion",
}

BOOLEAN_FIELDS = {
    "nivel_inicial_maternal", "nivel_inicial_infantes", "primario",
    "secundario", "adultos", "formacion_profesional", "alfabetizacion",
}


def _normalizar(texto: str) -> str:
    texto = texto.strip().lower()
    texto = "".join(c for c in unicodedata.normalize("NFKD", texto) if not unicodedata.combining(c))
    return texto


@router.get("/padron/estado", response_model=PadronEstadoResponse)
def estado_padron(db: Session = Depends(get_db)):
    total = db.query(EstablecimientoEstado).count()
    ultima = db.query(PadronImportacion).order_by(PadronImportacion.fecha.desc()).first()
    return PadronEstadoResponse(
        total_registros=total,
        ultima_importacion=ultima.fecha if ultima else None,
        ultimo_usuario_id=ultima.usuario_id if ultima else None,
        ultimo_total_procesados=ultima.total_procesados if ultima else None,
        ultimo_insertados=ultima.insertados if ultima else None,
        ultimo_actualizados=ultima.actualizados if ultima else None,
    )


@router.post("/padron/importar")
def importar_padron(
    file: UploadFile = File(...),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El archivo debe ser .xlsx")

    contenido = file.file.read()
    try:
        wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se pudo leer el archivo Excel")

    hoja = wb.active
    todas_las_filas = list(hoja.iter_rows(values_only=True))

    # Buscamos la fila de encabezados real (la que tiene "cueanexo") en vez de asumir
    # que es la fila 13 — el archivo real puede tener filas vacías/título al inicio
    # en cantidad distinta a la esperada.
    fila_encabezados_idx = None
    columna_a_campo: dict[int, str] = {}
    for i, fila in enumerate(todas_las_filas[:40]):
        if not fila:
            continue
        candidato: dict[int, str] = {}
        for idx, encabezado in enumerate(fila):
            if not encabezado:
                continue
            campo = ENCABEZADOS_PADRON.get(_normalizar(str(encabezado)))
            if campo:
                candidato[idx] = campo
        if "cueanexo" in candidato.values():
            fila_encabezados_idx = i
            columna_a_campo = candidato
            break

    if fila_encabezados_idx is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No se encontró una fila de encabezados con la columna 'cueanexo' en las primeras 40 filas",
        )

    existentes = {e.cueanexo: e for e in db.query(EstablecimientoEstado).all()}
    procesados = insertados = actualizados = 0
    hoy = date.today()

    for fila in todas_las_filas[fila_encabezados_idx + 1:]:
        if not fila:
            continue
        # Ojo: el archivo real repite nombres de columna (ej. "Primario" aparece una vez
        # por modalidad: Común, Especial, Adultos, etc.). Para los campos booleanos
        # combinamos esas columnas con OR en vez de dejar que la última pise a las anteriores.
        valores: dict = {}
        for idx, campo in columna_a_campo.items():
            if idx >= len(fila):
                continue
            valor = fila[idx]
            if campo in BOOLEAN_FIELDS:
                valores[campo] = bool(valores.get(campo)) or (
                    bool(valor) and str(valor).strip() not in ("0", "", "False", "NO", "No")
                )
            else:
                valores[campo] = valor

        cueanexo = valores.get("cueanexo")
        if not cueanexo:
            continue
        cueanexo = str(cueanexo).strip()
        procesados += 1

        if cueanexo in existentes:
            estab = existentes[cueanexo]
            for campo, valor in valores.items():
                setattr(estab, campo, valor)
            estab.actualizado_en = hoy
            actualizados += 1
        else:
            valores["cueanexo"] = cueanexo
            valores["actualizado_en"] = hoy
            estab = EstablecimientoEstado(**valores)
            db.add(estab)
            existentes[cueanexo] = estab
            insertados += 1

    registro = PadronImportacion(
        usuario_id=current_user.id,
        total_procesados=procesados,
        insertados=insertados,
        actualizados=actualizados,
    )
    db.add(registro)
    db.commit()

    return {
        "total_procesados": procesados,
        "insertados": insertados,
        "actualizados": actualizados,
    }


# --- Importación por lotes (el navegador parsea el Excel y manda filas en JSON) ---
# El archivo real del Ministerio supera el límite de payload de API Gateway (10 MB),
# así que /padron/importar (subida directa del .xlsx) solo funciona con archivos chicos.
# Para el padrón real, el frontend lee el Excel con SheetJS y manda los datos ya
# parseados en lotes a este endpoint.

class EstablecimientoBatchItem(BaseModel):
    cueanexo: str
    jurisdiccion: str | None = None
    sector: str | None = None
    ambito: str | None = None
    departamento: str | None = None
    cod_departamento: str | None = None
    localidad: str | None = None
    cod_localidad: str | None = None
    nombre: str | None = None
    domicilio: str | None = None
    codigo_postal: str | None = None
    telefono: str | None = None
    mail: str | None = None
    nivel_inicial_maternal: bool = False
    nivel_inicial_infantes: bool = False
    primario: bool = False
    secundario: bool = False
    adultos: bool = False
    formacion_profesional: bool = False
    alfabetizacion: bool = False


class ImportarBatchRequest(BaseModel):
    items: list[EstablecimientoBatchItem]


class ImportarBatchResponse(BaseModel):
    procesados: int
    insertados: int
    actualizados: int


@router.post("/padron/importar-batch", response_model=ImportarBatchResponse)
def importar_padron_batch(body: ImportarBatchRequest, db: Session = Depends(get_db)):
    """
    Upsert masivo en una sola sentencia SQL (INSERT ... ON DUPLICATE KEY UPDATE).
    Insertar fila por fila vía el ORM (db.add por cada item) era demasiado lento para
    lotes de mil+ filas contra TiDB — superaba el timeout de Lambda y el navegador
    reportaba "Failed to fetch" (la respuesta de error de API Gateway no trae headers
    CORS, así que el fetch se cae en vez de mostrar el error real).
    """
    if not body.items:
        return ImportarBatchResponse(procesados=0, insertados=0, actualizados=0)

    cueanexos = [item.cueanexo for item in body.items]
    existentes = set(
        row[0] for row in db.query(EstablecimientoEstado.cueanexo)
        .filter(EstablecimientoEstado.cueanexo.in_(cueanexos)).all()
    )

    hoy = date.today()
    filas = [{**item.model_dump(), "actualizado_en": hoy} for item in body.items]

    tabla = EstablecimientoEstado.__table__
    stmt = mysql_insert(tabla).values(filas)
    columnas_actualizables = [c.name for c in tabla.columns if c.name not in ("id", "cueanexo")]
    stmt = stmt.on_duplicate_key_update({col: stmt.inserted[col] for col in columnas_actualizables})
    db.execute(stmt)
    db.commit()

    insertados = sum(1 for c in cueanexos if c not in existentes)
    actualizados = len(body.items) - insertados
    return ImportarBatchResponse(procesados=len(body.items), insertados=insertados, actualizados=actualizados)


class RegistrarImportacionRequest(BaseModel):
    total_procesados: int
    insertados: int
    actualizados: int


@router.post("/padron/registrar-importacion")
def registrar_importacion(
    body: RegistrarImportacionRequest,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Se llama una vez al final, después de mandar todos los lotes, para dejar un registro de auditoría."""
    registro = PadronImportacion(usuario_id=current_user.id, **body.model_dump())
    db.add(registro)
    db.commit()
    return {"ok": True}


# ============================================================
# Descargas — informes cualitativos (Word) de Drive
# ============================================================

_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
_GDOC_MIME = "application/vnd.google-apps.document"
_FOLDER_MIME = "application/vnd.google-apps.folder"

_INVALID_FILENAME_CHARS = str.maketrans('\\/:*?"<>|', "-" * 9)


def _sanitize_filename(name: str) -> str:
    return name.translate(_INVALID_FILENAME_CHARS).strip()


def _nombre_con_prefijo(carpeta_padre: str, base: str) -> str:
    """
    Antepone el nombre de la carpeta contenedora al nombre del archivo para
    poder identificarlo sin ambigüedad, salvo que el archivo ya empiece con
    ese mismo nombre (frecuente: muchos EE ya suben el Word con su propio
    nombre adelante) — en ese caso no lo duplica.
    """
    if not carpeta_padre:
        return base.strip()
    if _normalizar(base).startswith(_normalizar(carpeta_padre)):
        return base.strip()
    return f"{carpeta_padre} - {base.strip()}"


def _listar_una_carpeta(args):
    """Lista el contenido directo de una carpeta con un cliente de Drive propio del hilo."""
    from scripts.scraper_control import build_services

    folder_id, carpeta_padre = args
    _, drive_hilo = build_services()
    archivos = []
    subcarpetas = []
    page_token = None
    while True:
        resp = drive_hilo.files().list(
            q=f"'{folder_id}' in parents and trashed = false",
            fields="nextPageToken, files(id, name, mimeType, modifiedTime)",
            includeItemsFromAllDrives=True,
            supportsAllDrives=True,
            pageToken=page_token,
            corpora="allDrives",
            pageSize=200,
        ).execute()
        for item in resp.get("files", []):
            if item["mimeType"] == _FOLDER_MIME:
                subcarpetas.append((item["id"], item["name"]))
            elif item["mimeType"] in (_DOCX_MIME, _GDOC_MIME):
                archivos.append((carpeta_padre, item["id"], item["name"], item["mimeType"], item.get("modifiedTime", "")))
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return archivos, subcarpetas


def _buscar_informes_word(folder_id: str):
    """
    Recorre recursivamente la carpeta (carpetas y subcarpetas, sin límite de
    profundidad) y devuelve todos los .docx y Documentos de Google que
    encuentra, junto con el nombre de la carpeta contenedora inmediata
    (puede ser una Diócesis, un Emaús, o vacío si el archivo está en la raíz)
    y su fecha de última modificación (RFC3339 UTC).

    Lista las carpetas de cada nivel en paralelo (cada carpeta = 1 llamada a
    la API de Drive; con ~50 Emaús hacerlo secuencial tarda ~20s) — cada
    hilo arma su propio cliente porque httplib2 no es thread-safe.
    """
    from concurrent.futures import ThreadPoolExecutor

    encontrados = []
    frontera = [(folder_id, "")]
    with ThreadPoolExecutor(max_workers=15) as ex:
        while frontera:
            resultados = list(ex.map(_listar_una_carpeta, frontera))
            siguiente = []
            for archivos, subcarpetas in resultados:
                encontrados.extend(archivos)
                siguiente.extend(subcarpetas)
            frontera = siguiente
    return encontrados


_FECHA_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


_SOFFICE_EXTRACT_LOCK = threading.Lock()
_LO_TAR_BR = "/opt/lo.tar.br"  # así viene la Layer pública "libreoffice-brotli" de shelfio
_LO_EXTRACTED_DIR = "/tmp/lo_libreoffice"
_LO_EXTRACTED_BIN = f"{_LO_EXTRACTED_DIR}/instdir/program/soffice.bin"


def _extraer_libreoffice_si_hace_falta() -> None:
    """
    La Layer pública "libreoffice-brotli" no trae el binario ya extraído:
    trae un único archivo /opt/lo.tar.br (tar comprimido con brotli, ~96MB
    -> ~370MB descomprimido) — Lambda no permite escribir en /opt, así que
    hay que descomprimirlo a /tmp una vez por contenedor. Se hace bajo un
    lock porque varios hilos de conversión pueden llamar a _soffice_bin()
    al mismo tiempo (paralelizado). Se cachea en /tmp: en una invocación
    "warm" (mismo contenedor reusado) no vuelve a descomprimir. Costo medido
    localmente: ~1.7s la primera vez.
    """
    if os.path.exists(_LO_EXTRACTED_BIN):
        return
    with _SOFFICE_EXTRACT_LOCK:
        if os.path.exists(_LO_EXTRACTED_BIN):  # doble chequeo: otro hilo pudo terminar mientras esperábamos el lock
            return
        import brotli
        import tarfile

        with open(_LO_TAR_BR, "rb") as f:
            datos = brotli.decompress(f.read())
        with tarfile.open(fileobj=io.BytesIO(datos)) as tf:
            tf.extractall(_LO_EXTRACTED_DIR)
        os.chmod(_LO_EXTRACTED_BIN, 0o755)


def _soffice_bin() -> str:
    """
    Ruta del binario de LibreOffice usado para convertir .docx a PDF.
    - Local (dev): "soffice" en el PATH (instalado con Homebrew/instalador oficial).
    - Lambda con la Layer pública "libreoffice-brotli": /opt/lo.tar.br, hay
      que descomprimirla a /tmp (ver _extraer_libreoffice_si_hace_falta()).
    - Cualquier otra Layer: si ya deja el binario extraído en /opt, se usa
      directo; si la ruta no es /opt/instdir/program/soffice.bin, overridear
      con la variable de entorno SOFFICE_BIN sin tocar código.
    """
    override = os.getenv("SOFFICE_BIN")
    if override:
        return override
    if os.path.exists(_LO_TAR_BR):
        _extraer_libreoffice_si_hace_falta()
        return _LO_EXTRACTED_BIN
    layer_bin = "/opt/instdir/program/soffice.bin"
    if os.path.exists(layer_bin):
        return layer_bin
    return "soffice"


def _convertir_docx_a_pdf_batch(rutas: list, outdir: str) -> str | None:
    """
    Convierte un lote de .docx a PDF en un solo proceso de LibreOffice (el
    arranque de soffice tiene un costo fijo de ~5-10s; convertir de a uno
    por archivo sería carísimo con ~20 archivos). Cada lote usa su propio
    perfil de usuario temporal para poder correr varios lotes en paralelo
    sin que se pisen entre sí. Devuelve un mensaje de error o None si salió bien.
    """
    import shutil
    import subprocess
    import tempfile

    perfil = tempfile.mkdtemp(prefix="lo_profile_")
    try:
        resultado = subprocess.run(
            [
                _soffice_bin(),
                f"-env:UserInstallation=file://{perfil}",
                "--headless", "--norestore",
                "--convert-to", "pdf",
                "--outdir", outdir,
            ] + rutas,
            capture_output=True, text=True, timeout=120,
        )
        if resultado.returncode != 0:
            return resultado.stderr[-500:] or "soffice terminó con error sin detalle"
        return None
    except Exception as e:
        return str(e)
    finally:
        # Si no se limpia, estos perfiles de LibreOffice se acumulan en /tmp
        # durante toda la vida del contenedor de Lambda (se reutiliza entre
        # invocaciones) hasta agotar el espacio efímero disponible.
        shutil.rmtree(perfil, ignore_errors=True)


def _convertir_docx_a_pdf_paralelo(items: list, outdir: str):
    """
    items: lista de (clave, ruta_docx). Divide en lotes y corre varios
    procesos de soffice en paralelo (cada uno con su perfil propio).
    Devuelve (resultado, errores_lotes): resultado es {clave: ruta_pdf o
    None si falló} y errores_lotes es la lista de stderr de los lotes que
    fallaron (para diagnóstico, un lote agrupa varios archivos).
    """
    from concurrent.futures import ThreadPoolExecutor

    if not items:
        return {}, []

    n_lotes = min(6, max(1, len(items) // 4 or 1))
    lotes = [items[i::n_lotes] for i in range(n_lotes)]
    lotes = [l for l in lotes if l]

    def _procesar_lote(lote):
        rutas = [ruta for _clave, ruta in lote]
        error = _convertir_docx_a_pdf_batch(rutas, outdir)
        salida = {}
        for clave, ruta in lote:
            pdf_path = os.path.splitext(ruta)[0] + ".pdf"
            salida[clave] = pdf_path if (error is None and os.path.exists(pdf_path)) else None
        return salida, error

    resultado = {}
    errores_lotes = []
    with ThreadPoolExecutor(max_workers=n_lotes) as ex:
        for salida, error in ex.map(_procesar_lote, lotes):
            resultado.update(salida)
            if error:
                errores_lotes.append(error)
    return resultado, errores_lotes


def _generar_pagina_divisoria(nombre_emaus: str) -> bytes:
    """
    Página simple con el nombre del Emaús en grande, para intercalar antes
    de cada informe en el PDF combinado. No depende de si el título interno
    del documento original menciona o no el Emaús — se aplica igual a todos.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    c.setFont("Helvetica-Bold", 28)
    c.drawCentredString(width / 2, height / 2, nombre_emaus)
    c.setFont("Helvetica", 12)
    c.drawCentredString(width / 2, height / 2 - 1.2 * cm, "Informe cualitativo — Medio Término 2026")
    c.showPage()
    c.save()
    return buf.getvalue()


def _merge_pdfs(items: list) -> bytes:
    """
    items: lista de (nombre_emaus, ruta_o_bytes_del_pdf). Antepone una
    página divisoria con el nombre del Emaús antes de las páginas de cada
    informe.
    """
    from pypdf import PdfWriter

    writer = PdfWriter()
    for nombre_emaus, fuente in items:
        writer.append(io.BytesIO(_generar_pagina_divisoria(nombre_emaus)))
        if isinstance(fuente, (bytes, bytearray)):
            writer.append(io.BytesIO(fuente))
        else:
            writer.append(fuente)
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


@router.get("/informes-cualitativos/zip")
def descargar_informes_cualitativos(
    modificados_desde: str | None = None,
    incluir_pdf_combinado: bool = False,
    db: Session = Depends(get_db),
):
    """
    Recorre la carpeta DRIVE_FOLDER_ID (carpetas y subcarpetas de Diócesis/
    Emaús) y arma un único ZIP con todos los informes cualitativos en
    formato Word: tanto los .docx ya subidos como los Documentos de Google
    (se convierten a .docx al vuelo vía la API de exportación de Drive).

    El ZIP no se devuelve directo en la respuesta: Lambda tiene un límite
    duro de 6MB para respuestas síncronas, muy por debajo de lo que pesa
    este ZIP. Se sube a S3 (bucket en S3_BUCKET_DESCARGAS) y se devuelve
    un JSON con una URL prefirmada de descarga (vence en 10 minutos).

    Solo se incluyen archivos cuya carpeta contenedora inmediata coincide
    con el nombre de un Emaús real (excluye archivos sueltos en la raíz de
    Drive y carpetas de agrupación que no son Emaús).

    Si se pasa `modificados_desde` (fecha YYYY-MM-DD), solo se incluyen los
    archivos modificados en o después de esa fecha (inclusive).

    Cada archivo se nombra "<carpeta contenedora> - <nombre original>.docx"
    para poder identificarlos sin ambigüedad, ya que los nombres no son
    consistentes entre carpetas (salvo que el nombre ya empiece con el de
    la carpeta, para no duplicarlo).

    Si `incluir_pdf_combinado=True`, además arma un PDF único con todas las
    páginas de todos los informes (en el mismo orden que aparecen en el
    ZIP) y lo agrega como archivo adicional dentro del ZIP. Los Documentos
    de Google se exportan a PDF directo vía Drive; los .docx reales se
    convierten con LibreOffice (requiere el binario disponible — ver
    _soffice_bin()). Es opt-in porque agrega ~15-20s extra a la respuesta.

    Las descargas se hacen en paralelo (cada hilo arma su propio cliente de
    Drive — httplib2, usado por la librería de Google, no es thread-safe si
    se comparte una sola instancia entre hilos) para que ~45 archivos no
    tarden más de un minuto y arriesguen el timeout de API Gateway; medido:
    ~80s en secuencial vs ~10s en paralelo con 10 workers.
    """
    import tempfile
    import zipfile
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from scripts.scraper_control import build_services

    if modificados_desde and not _FECHA_RE.match(modificados_desde):
        raise HTTPException(status_code=400, detail="modificados_desde debe tener formato YYYY-MM-DD")

    folder_id = os.getenv("DRIVE_FOLDER_ID", "")
    if not folder_id:
        raise HTTPException(status_code=500, detail="DRIVE_FOLDER_ID no configurado")

    # La consulta de Emaús corre en paralelo con el recorrido de Drive (son
    # independientes) — usa su propia sesión de DB porque `db` es la sesión
    # inyectada por request y SQLAlchemy no permite compartir una sesión
    # entre hilos.
    def _cargar_nombres_emaus():
        from app.database import SessionLocal
        db_hilo = SessionLocal()
        try:
            return {_normalizar(e.nombre) for e in db_hilo.query(Emaus).all()}
        finally:
            db_hilo.close()

    with ThreadPoolExecutor(max_workers=2) as setup_ex:
        fut_nombres = setup_ex.submit(_cargar_nombres_emaus)
        fut_encontrados = setup_ex.submit(_buscar_informes_word, folder_id)
        nombres_emaus = fut_nombres.result()
        encontrados = fut_encontrados.result()

    # Solo carpetas que son Emaús reales — descarta la raíz y carpetas de agrupación
    encontrados = [f for f in encontrados if f[0] and _normalizar(f[0]) in nombres_emaus]

    # Filtro opcional por fecha de modificación (inclusive) — comparación de
    # strings RFC3339 UTC, válida porque el formato es de ancho fijo
    if modificados_desde:
        desde_ts = f"{modificados_desde}T00:00:00"
        encontrados = [f for f in encontrados if f[4] and f[4] >= desde_ts]

    if not encontrados:
        raise HTTPException(status_code=404, detail="No se encontraron informes Word con esos filtros")

    # Se procesan en dos carriles independientes que corren en paralelo entre
    # sí (no solo cada uno internamente): los .docx reales necesitan pasar
    # por LibreOffice (lento), los Documentos de Google exportan a PDF
    # directo por Drive (rápido). Separarlos evita que la conversión con
    # LibreOffice espere a que terminen de bajar los Documentos de Google,
    # ganando ~8-10s medidos contra Drive real — necesario para no arriesgar
    # el timeout de API Gateway cuando incluir_pdf_combinado=True.
    docx_items = [it for it in encontrados if it[3] == _DOCX_MIME]
    gdoc_items = [it for it in encontrados if it[3] != _DOCX_MIME]

    def _descargar_docx_raw(item):
        carpeta_padre, file_id, nombre, mime_type, _modified = item
        try:
            _, drive_hilo = build_services()
            contenido = drive_hilo.files().get_media(fileId=file_id, supportsAllDrives=True).execute()
            base = nombre if nombre.lower().endswith(".docx") else f"{nombre}.docx"
            return {
                "carpeta_padre": carpeta_padre, "base": base, "contenido": contenido,
                "mime_type": mime_type, "pdf_bytes": None, "error": None,
            }
        except Exception as e:
            return {
                "carpeta_padre": carpeta_padre, "base": nombre, "contenido": None,
                "mime_type": mime_type, "pdf_bytes": None, "error": str(e),
            }

    def _descargar_gdoc(item):
        carpeta_padre, file_id, nombre, mime_type, _modified = item
        try:
            _, drive_hilo = build_services()
            contenido = drive_hilo.files().export_media(fileId=file_id, mimeType=_DOCX_MIME).execute()
        except Exception as e:
            return {
                "carpeta_padre": carpeta_padre, "base": nombre, "contenido": None,
                "mime_type": mime_type, "pdf_bytes": None, "error": str(e),
            }

        pdf_bytes = None
        pdf_error = None
        if incluir_pdf_combinado:
            try:
                _, drive_hilo2 = build_services()
                pdf_bytes = drive_hilo2.files().export_media(fileId=file_id, mimeType="application/pdf").execute()
            except Exception as e:
                # El .docx ya se descargó bien — que falle solo el export a PDF
                # no debe sacar el archivo del ZIP, solo del PDF combinado.
                pdf_error = str(e)

        return {
            "carpeta_padre": carpeta_padre, "base": f"{nombre}.docx", "contenido": contenido,
            "mime_type": mime_type, "pdf_bytes": pdf_bytes, "error": None, "pdf_error": pdf_error,
        }

    errores = []
    tmpdir_ctx = tempfile.TemporaryDirectory(prefix="informes_pdf_")
    try:
        tmpdir = tmpdir_ctx.name

        def _carril_docx():
            """Baja los .docx reales y, si corresponde, los convierte a PDF con LibreOffice."""
            locales = []
            with ThreadPoolExecutor(max_workers=10) as ex:
                for r in ex.map(_descargar_docx_raw, docx_items):
                    locales.append(r)

            errores_conversion = []
            if incluir_pdf_combinado:
                pendientes = []
                for idx, r in enumerate(locales):
                    if r["error"]:
                        continue
                    ruta = os.path.join(tmpdir, f"docx_{idx}.docx")
                    with open(ruta, "wb") as f:
                        f.write(r["contenido"])
                    pendientes.append((idx, ruta))

                rutas_pdf_por_idx, errores_conversion = _convertir_docx_a_pdf_paralelo(pendientes, tmpdir)
                for idx, ruta_pdf in rutas_pdf_por_idx.items():
                    if ruta_pdf:
                        locales[idx]["pdf_ruta"] = ruta_pdf
                    else:
                        errores_conversion.append(f"{locales[idx]['carpeta_padre']} / {locales[idx]['base']}: no se pudo convertir a PDF")
            return locales, errores_conversion

        def _carril_gdocs():
            """Exporta los Documentos de Google a .docx (y a PDF directo si corresponde)."""
            locales = []
            with ThreadPoolExecutor(max_workers=15) as ex:
                for r in ex.map(_descargar_gdoc, gdoc_items):
                    locales.append(r)
            return locales

        with ThreadPoolExecutor(max_workers=2) as carriles_ex:
            fut_docx = carriles_ex.submit(_carril_docx)
            fut_gdocs = carriles_ex.submit(_carril_gdocs)
            resultados_docx, errores_conversion = fut_docx.result()
            resultados_gdocs = fut_gdocs.result()

        resultados = resultados_docx + resultados_gdocs
        errores.extend(errores_conversion)

        buffer = io.BytesIO()
        usados = set()

        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            entradas_ordenadas = []
            for r in resultados:
                if r["error"]:
                    errores.append(f"{r['carpeta_padre'] or '(raíz)'} / {r['base']}: {r['error']}")
                    continue
                zip_name = _sanitize_filename(_nombre_con_prefijo(r["carpeta_padre"], r["base"]))
                final_name = zip_name
                i = 2
                while final_name in usados:
                    stem, ext = os.path.splitext(zip_name)
                    final_name = f"{stem} ({i}){ext}"
                    i += 1
                usados.add(final_name)
                zf.writestr(final_name, r["contenido"])
                entradas_ordenadas.append((final_name, r))

            if incluir_pdf_combinado:
                # Mismo orden que los archivos individuales, para que el PDF
                # combinado se lea en el mismo orden en que aparecen en el ZIP
                entradas_ordenadas.sort(key=lambda t: t[0])
                fuentes_pdf = []
                for _final_name, r in entradas_ordenadas:
                    if r.get("pdf_bytes"):
                        fuentes_pdf.append((r["carpeta_padre"], r["pdf_bytes"]))
                    elif r.get("pdf_ruta"):
                        fuentes_pdf.append((r["carpeta_padre"], r["pdf_ruta"]))
                    elif r.get("pdf_error"):
                        errores.append(f"{r['carpeta_padre']} / {r['base']}: no se pudo exportar a PDF para el combinado ({r['pdf_error']}) — sí está en el ZIP")
                if fuentes_pdf:
                    try:
                        pdf_combinado = _merge_pdfs(fuentes_pdf)
                        hoy_nombre = datetime.now().strftime("%Y%m%d")
                        sufijo = f"_modificados_{modificados_desde.replace('-', '')}" if modificados_desde else ""
                        zf.writestr(f"TODOS_LOS_INFORMES_{hoy_nombre}{sufijo}.pdf", pdf_combinado)
                    except Exception as e:
                        errores.append(f"No se pudo armar el PDF combinado: {e}")

            if errores:
                zf.writestr(
                    "_ERRORES.txt",
                    "No se pudieron descargar/convertir los siguientes archivos:\n\n" + "\n".join(errores),
                )
    finally:
        tmpdir_ctx.cleanup()

    hoy = datetime.now().strftime("%Y%m%d")
    if modificados_desde:
        nombre_zip = f"informes_cualitativos_{hoy}_modificados_{modificados_desde.replace('-', '')}.zip"
    else:
        nombre_zip = f"informes_cualitativos_{hoy}.zip"

    # Lambda no puede devolver más de 6MB en una respuesta síncrona (límite
    # duro, no configurable) — con ~44 informes el ZIP supera eso by lejos.
    # Por eso se sube a S3 y se devuelve un link temporal en vez del archivo.
    s3_bucket = os.getenv("S3_BUCKET_DESCARGAS", "")
    if not s3_bucket:
        raise HTTPException(status_code=500, detail="S3_BUCKET_DESCARGAS no configurado")

    import uuid
    import boto3
    from botocore.client import Config

    # signature_version explícito: sin esto, boto3 firma con el formato
    # viejo (SigV2) en us-east-1 por compatibilidad histórica — muchas
    # cuentas de AWS lo tienen deshabilitado y S3 rechaza la URL con 403
    # (el navegador termina "descargando" ese error XML chico en vez del ZIP).
    s3 = boto3.client(
        "s3",
        region_name=os.getenv("AWS_REGION", "us-east-1"),
        config=Config(signature_version="s3v4"),
    )
    key = f"informes-cualitativos/{uuid.uuid4().hex}.zip"
    buffer.seek(0)
    s3.put_object(
        Bucket=s3_bucket,
        Key=key,
        Body=buffer.getvalue(),
        ContentType="application/zip",
    )
    url = s3.generate_presigned_url(
        "get_object",
        Params={
            "Bucket": s3_bucket,
            "Key": key,
            "ResponseContentDisposition": f'attachment; filename="{nombre_zip}"',
            "ResponseContentType": "application/zip",
        },
        ExpiresIn=600,  # 10 minutos — el bucket también autoborra el objeto a 1 día
    )
    return {"download_url": url, "filename": nombre_zip}
